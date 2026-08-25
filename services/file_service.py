# -*- coding: utf-8 -*-
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from keyboards.admin import export_status_labels


def build_export_file(tickets):
    wb = Workbook()
    ws = wb.active
    ws.title = "Murojaatlar"

    headers = [
        "№", "Sana", "Foydalanuvchi", "Telegram username", "Telegram ID", "Telefon",
        "Bo'lim", "Muammo turi", "Tavsif",
        "Holat", "Mas'ul", "Izoh (yechim)", "Yopilgan vaqt",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_labels = export_status_labels()
    for ticket in tickets:
        ws.append([
            ticket["id"],
            ticket["created_at"],
            ticket["department"],
            ticket["user_name"],
            ticket.get("telegram_username") or "",
            ticket["user_id"],
            ticket.get("phone_number") or "",
            ticket["category"],
            ticket["description"],
            status_labels.get(ticket["status"], ticket["status"]),
            ticket.get("assigned_to_name") or "",
            ticket.get("resolution_comment") or "",
            ticket.get("closed_at") or "",
        ])

    widths = [6, 18, 22, 22, 16, 18, 22, 28, 40, 14, 20, 40, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"murojaatlar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return filename, buffer.read()
